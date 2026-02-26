"""
Service for generating professional Excel reports for customer analysis.
Adapted from external best practices for maximum utility.
"""
import pandas as pd
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

class ReportService:
    """Service to generate detailed Excel reports"""
    
    def generate_excel_report(self, customer_name, profile_data, installed_base_df=None):
        """
        Generate a comprehensive multi-tab Excel report
        
        Args:
            customer_name: Name of the customer
            profile_data: Dictionary containing the AI-generated profile (Steckbrief)
            installed_base_df: Optional DataFrame of installed base for detailed tab
            
        Returns:
            BytesIO buffer containing the Excel file
        """
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Define styles
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        title_font = Font(name='Arial', size=14, bold=True)
        subtitle_font = Font(name='Arial', size=11, bold=True)
        cell_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # --- TAB 1: CUSTOMER OVERVIEW ---
        ws_overview = wb.create_sheet("Customer Overview")
        ws_overview.cell(row=1, column=1, value='Customer Analysis Report').font = title_font
        ws_overview.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        # Basic Data
        ws_overview.cell(row=4, column=1, value='Company Profile').font = subtitle_font
        
        basic = profile_data.get('basic_data', {})
        basic_fields = [
            ('Company Name', basic.get('name')),
            ('Headquarters', basic.get('hq_address')),
            ('Owner/Parent', basic.get('owner')),
            ('CEO', basic.get('ceo')),
            ('Employees (FTE)', basic.get('fte')),
            ('Financial Status', basic.get('financials')),
            ('Company Focus', basic.get('company_focus'))
        ]
        
        current_row = 5
        for label, value in basic_fields:
            ws_overview.cell(row=current_row, column=1, value=label).font = Font(bold=True)
            ws_overview.cell(row=current_row, column=2, value=str(value or '—'))
            current_row += 1
            
        ws_overview.column_dimensions['A'].width = 25
        ws_overview.column_dimensions['B'].width = 60
        
        # --- TAB 2: MARKET INTELLIGENCE ---
        ws_intel = wb.create_sheet("Market Intelligence")
        ws_intel.cell(row=1, column=1, value='Market Intelligence & Strategy').font = title_font
        
        intel = profile_data.get('market_intelligence', {})
        strategy = profile_data.get('sales_strategy', {})
        metallurgy = profile_data.get('metallurgical_insights', {})
        
        sections = [
            ('💰 Financial Health', intel.get('financial_health')),
            ('📰 Recent Developments', intel.get('recent_developments')),
            ('📊 Market Position', intel.get('market_position')),
            ('🎯 Strategic Outlook', intel.get('strategic_outlook')),
            ('⚠️ Risk Assessment', intel.get('risk_assessment')),
            ('🛠️ Technical Insights', metallurgy.get('modernization_potential')),
            ('💡 Sales Strategy', strategy.get('suggested_next_steps'))
        ]
        
        current_row = 3
        for title, content in sections:
            if content:
                # Header
                cell = ws_intel.cell(row=current_row, column=1, value=title)
                cell.font = subtitle_font
                cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
                ws_intel.merge_cells(f'A{current_row}:B{current_row}')
                current_row += 1
                
                # Content
                cell = ws_intel.cell(row=current_row, column=1, value=str(content))
                cell.alignment = cell_alignment
                ws_intel.merge_cells(f'A{current_row}:B{current_row}')
                
                # Estimate height based on text length (rough approximation)
                lines = len(str(content)) // 100 + 2
                ws_intel.row_dimensions[current_row].height = min(lines * 15, 200)
                
                current_row += 2
        
        ws_intel.column_dimensions['A'].width = 15
        ws_intel.column_dimensions['B'].width = 80
        
        # --- TAB 3: INSTALLED BASE ---
        if installed_base_df is not None and not installed_base_df.empty:
            ws_ib = wb.create_sheet("Installed Base")
            ws_ib.cell(row=1, column=1, value='Installed Base Inventory').font = title_font
            
            # Write header
            cols = list(installed_base_df.columns)
            for c_idx, col_name in enumerate(cols, 1):
                cell = ws_ib.cell(row=3, column=c_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                ws_ib.column_dimensions[get_column_letter(c_idx)].width = 20
            
            # Write data
            for r_idx, row in enumerate(dataframe_to_rows(installed_base_df, index=False, header=False), 4):
                for c_idx, value in enumerate(row, 1):
                    cell = ws_ib.cell(row=r_idx, column=c_idx, value=value)
                    cell.alignment = Alignment(horizontal='left')
            
            # Add table
            try:
                end_col = get_column_letter(len(cols))
                end_row = 3 + len(installed_base_df)
                tab = Table(displayName="InstalledBase", ref=f"A3:{end_col}{end_row}")
                style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
                tab.tableStyleInfo = style
                ws_ib.add_table(tab)
            except Exception:
                pass

        # Save
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

report_service = ReportService()
